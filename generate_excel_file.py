import xlsxwriter
from openpyxl import load_workbook
import ast
from uuid import uuid4
import random
from fashion.models.item import Item

class ExcelTemplateGenerator:


    @staticmethod
    def get_data(fn: str):
        wb = load_workbook(fn)
        sheet = wb.get_sheet_by_name('items')
        items = []
        sizes = []
        user_id =f'{uuid4()}'
        for x in range(2, sheet.max_row):
            iid = f'{uuid4()}'
            items.append([iid, sheet.cell(row=x, column=1).value, sheet.cell(row=x, column=2).value, sheet.cell(row=x, column=3).value,
                          sheet.cell(row=x, column=6).value, sheet.cell(row=x, column=7).value, sheet.cell(row=x, column=8).value, sheet.cell(row=x, column=9).value,
                          sheet.cell(row=x, column=10).value, sheet.cell(row=x, column=11).value, sheet.cell(row=x, column=13).value, user_id])
            for s in ast.literal_eval(sheet.cell(row=x, column=5).value):
                sizes.append([f'{uuid4()}', s, random.choice(range(1, 10)), iid])
        users=[[user_id,'hadi','daher','hadi_daher@fashion.com','hadi123','hadi123','Admin','tripoli','+96171713662']]
        return items, sizes, users

    @staticmethod
    def generate_excel_template(filename: str):
        workbook = xlsxwriter.Workbook(filename)
        items_prop = [
            {'header': 'iid', 'type': str},
            {'header': 'item_name', 'type': str},
            {'header': 'item_category', 'type': str},
            {'header': 'gender', 'type': str},
            {'header': 'flash_sale', 'type': int},
            {'header': 'actual_price', 'type': int},
            {'header': 'currency', 'type': str},
            {'header': 'item_model', 'type': str},
            {'header': 'reference', 'type': str},
            {'header': 'status', 'type': str},
            {'header': 'description', 'type': str},
            {'header': 'user_id', 'type': str}
        ]
        sizes_prop = [
            {'header': 'iid', 'type': str},
            {'header': 'size', 'type': str},
            {'header': 'quantity', 'type': int},
            {'header': 'item_id', 'type': Item}
        ]
  
        users_prop = [
            {'header': 'iid', 'type':str},
            {'header': 'first_name', 'type':str},
            {'header': 'last_name', 'type':str},
            {'header': 'email', 'type':str},
            {'header': 'password', 'type':str},
            {'header': 'confirm_pass', 'type':str},
            {'header': 'user_role', 'type':str},
            {'header': 'adress', 'type':str},
            {'header': 'phone', 'type':str}
        ]
        sheet1 = workbook.add_worksheet('items')
        sheet2 = workbook.add_worksheet('sizes')
        sheet3 = workbook.add_worksheet('users')
        data1, data2, data3= ExcelTemplateGenerator.get_data('fashion/data/tmp/items_template.xlsx')
        opts1 = []
        options1 = {}
        opts2 = []
        options2 = {}
        opts3 =[]
        options3 = {}
        for prop in items_prop:
            opts1.append({'header': prop['header']})
            column = len(opts1)
            cell = chr(ord('A') + column - 1) + '1'
            if prop['type'] == str:
                sheet1.data_validation(0, column - 1, 1, column - 1, {'validate': 'custom', 'value': f'=ISTEXT({cell})'})
            elif prop['type'] in [int, float]:
                sheet1.data_validation(0, column - 1, 1, column - 1, {'validate': 'custom', 'value': f'=ISNUMBER({cell})'})
            options1['columns'] = opts1
            options1['data'] = data1
        sheet1.add_table(0, 0, len(data1), len(options1['columns']) - 1, options1)
        # --------------------------------------------------------------------------------------
        for prop in sizes_prop:
            opts2.append({'header': prop['header']})
            column = len(opts2)
            cell = chr(ord('A') + column - 1) + '1'
            if prop['type'] == str:
                sheet2.data_validation(0, column - 1, 1, column - 1, {'validate': 'custom', 'value': f'=ISTEXT({cell})'})
            elif prop['type'] in [int, float]:
                sheet2.data_validation(0, column - 1, 1, column - 1, {'validate': 'custom', 'value': f'=ISNUMBER({cell})'})
            options2['columns'] = opts2
            options2['data'] = data2
        sheet2.add_table(0, 0, len(data2), len(options2['columns']) - 1, options2)
        # --------------------------------------------------------------------------------------
        for prop in users_prop:
            opts3.append({'header': prop['header']})
            column = len(opts3)
            cell = chr(ord('A') + column - 1) + '1'
            if prop['type'] == str:
                sheet3.data_validation(0, column - 1, 1, column - 1, {'validate': 'custom', 'value': f'=ISTEXT({cell})'})
            elif prop['type'] in [int, float]:
                sheet3.data_validation(0, column - 1, 1, column - 1, {'validate': 'custom', 'value': f'=ISNUMBER({cell})'})
            options3['columns'] = opts3
            options3['data'] = data3
        sheet3.add_table(0, 0, len(data3), len(options3['columns']) - 1, options3)

        workbook.close()
        return workbook
    

ExcelTemplateGenerator.generate_excel_template('fashion/data/tmp/fashion_template.xlsx')
